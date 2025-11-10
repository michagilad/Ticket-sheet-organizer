#!/usr/bin/env python3
"""
Ticket Organizer Web Application
A Flask web app with a UI for uploading CSV files and downloading organized XLSX output.
"""

import os
import tempfile
from pathlib import Path
from datetime import datetime
from flask import Flask, render_template, request, send_file, flash, redirect, url_for, session, jsonify, after_this_request
from werkzeug.utils import secure_filename
import csv
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import gspread
from google.oauth2.service_account import Credentials

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'your-secret-key-change-in-production-please-use-random-string')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['UPLOAD_FOLDER'] = tempfile.gettempdir()

# Force logging to INFO level
import logging
logging.basicConfig(level=logging.INFO)
app.logger.setLevel(logging.INFO)

ALLOWED_EXTENSIONS = {'csv'}

# Available assignees list
ASSIGNEES = [
    'Cs. Zsófi',
    'Ricsi',
    'Attila',
    'Lili',
    'Flóra',
    'Csabi',
    'Marci',
    'Gábor',
    'Bálint',
    'Vivien',
    'B.Zsófi',
    'Muki',
    'Péter',
    'Adri',
    'Era',
    'Endre',
    'Vencel',
    'Áron'
]

# Google Sheets Configuration
GOOGLE_SHEETS_ENABLED = os.environ.get('GOOGLE_SHEETS_ENABLED', 'false').lower() == 'true'
GOOGLE_SHEET_NAME = os.environ.get('GOOGLE_SHEET_NAME', 'Tickets with contact')


def get_google_sheet_client():
    """Initialize and return Google Sheets client."""
    if not GOOGLE_SHEETS_ENABLED:
        return None
    
    try:
        # Load credentials from environment variable (JSON string)
        creds_json = os.environ.get('GOOGLE_SHEETS_CREDENTIALS')
        if not creds_json:
            app.logger.warning('GOOGLE_SHEETS_CREDENTIALS not found')
            return None
        
        # Parse JSON credentials
        creds_dict = json.loads(creds_json)
        
        # Define scopes
        scopes = [
            'https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/drive'
        ]
        
        # Create credentials
        credentials = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        
        # Create client
        client = gspread.authorize(credentials)
        return client
    except Exception as e:
        app.logger.error(f'Error initializing Google Sheets client: {e}', exc_info=True)
        return None


def get_existing_assignments_from_sheet():
    """
    Read ALL tabs in Google Sheet and return dict of Experience ID → Assignee mappings.
    Returns: dict {experience_id: {'assignee': str, 'experience_name': str, 'date_assigned': str}}
    """
    client = get_google_sheet_client()
    if not client:
        app.logger.info('Google Sheets not enabled, skipping history check')
        return {}
    
    try:
        # Open the spreadsheet
        spreadsheet = client.open(GOOGLE_SHEET_NAME)
        
        assignments = {}
        
        # Read all worksheets/tabs
        for worksheet in spreadsheet.worksheets():
            try:
                app.logger.info(f'Reading worksheet: {worksheet.title}')
                # Get all records (assumes first row is header)
                records = worksheet.get_all_records()
                app.logger.info(f'  Found {len(records)} rows in {worksheet.title}')
                
                for record in records:
                    # Use your column names
                    exp_id = str(record.get('Associated Experience IDs', '')).strip()
                    if exp_id and exp_id != '':
                        # Only store if not already stored (first occurrence wins)
                        if exp_id not in assignments:
                            assignments[exp_id] = {
                                'assignee': record.get('Assignee', ''),
                                'experience_name': record.get('Associated Experience', ''),
                                'date_assigned': record.get('Create date', ''),
                                'tab': worksheet.title
                            }
            except Exception as e:
                app.logger.warning(f'Error reading worksheet {worksheet.title}: {e}')
                continue
        
        app.logger.info(f'Loaded {len(assignments)} existing assignments from Google Sheets')
        return assignments
        
    except Exception as e:
        app.logger.error(f'Error reading Google Sheet: {e}', exc_info=True)
        return {}


def allowed_file(filename):
    """Check if file has allowed extension."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def read_csv_robust(file_path):
    """Read CSV file with robust handling of line breaks and special characters."""
    with open(file_path, 'r', encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    return rows


def extract_and_reorder_columns(rows):
    """Extract only the required columns and reorder them as specified."""
    # Define the columns we need from input
    input_columns = [
        'Ticket ID',
        'Backstage Experience page',
        'Company',
        'Create date',
        'Priority',
        'Record source detail 1',
        'Ticket description',
        'Ticket name',
        'Ticket status',
        'Associated Experience IDs',
        'Associated Experience',
        'Assignee',
        'Invalid ticket',
        'Ticket Category',
        'Sub Category',
        'Non-ticketed issues',
        'Additional Notes',
        'Notes - Admin',
        'Next step',
        'Last Updated - Status',
        'Last Updated - Next step',
        'Last Updated - Invalid Ticket',
        'Last Updated - Non Ticketed Issues',
        'Public Preview Link'  # This will be generated, not from input
    ]
    
    # Define the output order
    output_order = [
        'Associated Experience',
        'Backstage Experience page',
        'Public Preview Link',
        'Associated Experience IDs',
        'Assignee',
        'Ticket name',
        'Ticket description',
        'Ticket status',
        'Invalid ticket',
        'Ticket Category',
        'Sub Category',
        'Non-ticketed issues',
        'Additional Notes',
        'Notes - Admin',
        'Next step',
        'Create date',
        'Ticket ID',
        'Record source detail 1',
        'Company',
        'Priority',
        'Last Updated - Status',
        'Last Updated - Next step',
        'Last Updated - Invalid Ticket',
        'Last Updated - Non Ticketed Issues'
    ]
    
    processed_rows = []
    for row in rows:
        processed_row = {}
        for col in output_order:
            processed_row[col] = row.get(col, '')
        
        # If Associated Experience is empty, use Experience Name as fallback
        if not processed_row.get('Associated Experience') or not str(processed_row['Associated Experience']).strip():
            experience_name = row.get('Experience Name', '')
            if experience_name:
                processed_row['Associated Experience'] = experience_name
        
        processed_rows.append(processed_row)
    
    # Sort by Associated Experience IDs (numerically) before grouping
    # This ensures that all rows with the same Experience ID are consecutive
    def get_experience_id(row):
        exp_id = row.get('Associated Experience IDs', '')
        if not exp_id or not str(exp_id).strip():
            # Empty IDs go to the end
            return float('inf')
        try:
            # Try to convert to integer for proper numeric sorting
            return int(exp_id)
        except (ValueError, TypeError):
            # If conversion fails, put at the end
            return float('inf')
    
    processed_rows.sort(key=get_experience_id)
    
    return processed_rows, output_order


def assign_experiences_to_assignees(groups, assignee_distribution, returning_exp_map=None):
    """
    Assign experiences to assignees based on the distribution settings.
    
    Args:
        groups: List of experience groups
        assignee_distribution: Dict with assignee names as keys and counts as values
        returning_exp_map: Dict of returning experience IDs (to skip)
    
    Returns:
        List of groups with assignees assigned
    """
    # Filter out assignees with 0 count
    active_assignees = {name: count for name, count in assignee_distribution.items() if count > 0}
    
    if not active_assignees:
        return groups  # No assignees to assign
    
    # Create a list of assignees repeated by their count
    assignee_pool = []
    for name, count in active_assignees.items():
        assignee_pool.extend([name] * count)
    
    # Assign to groups (only those with non-empty experience IDs and NOT already assigned)
    assignee_idx = 0
    for group_key, group_size, group_rows in groups:
        # Check if this group has a valid experience ID
        exp_id = group_key[2]  # Associated Experience IDs is the third element
        has_exp_id = exp_id and str(exp_id).strip()
        
        # Skip if already assigned (returning experience)
        is_already_assigned = returning_exp_map and str(exp_id).strip() in returning_exp_map
        
        if has_exp_id and assignee_pool and not is_already_assigned:
            # Assign the same assignee to all rows in this group
            assignee = assignee_pool[assignee_idx % len(assignee_pool)]
            for row in group_rows:
                row['Assignee'] = assignee
            assignee_idx += 1
    
    return groups


def group_by_experience(rows):
    """Group rows by experience to identify which rows belong to the same experience."""
    groups = []
    current_group = []
    current_key = None
    
    for row in rows:
        # Create a key based on the three identifying columns
        key = (
            row.get('Associated Experience', ''),
            row.get('Backstage Experience page', ''),
            row.get('Associated Experience IDs', '')
        )
        
        # Check if key is all empty (don't group empty rows together)
        is_empty_key = all(k == '' for k in key)
        
        if current_key is None:
            # First row
            current_key = key
            current_group = [row]
        elif key == current_key and not is_empty_key:
            # Same experience AND not empty, add to current group
            current_group.append(row)
        else:
            # New experience OR empty key, save previous group and start new one
            groups.append((current_key, len(current_group), current_group))
            current_key = key
            current_group = [row]
    
    # Don't forget the last group
    if current_group:
        groups.append((current_key, len(current_group), current_group))
    
    return groups


def create_xlsx(rows, column_order, output_path):
    """Create XLSX file with proper formatting, merging cells for identical experiences."""
    app.logger.info('  Creating workbook...')
    wb = Workbook()
    ws = wb.active
    ws.title = "Organized Tickets"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Alternating colors for experience groups
    colors = [
        PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"),
        PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
    ]
    
    # Border styles
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    thick_border = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )
    
    # Write header row
    app.logger.info('  Writing header row...')
    for col_idx, col_name in enumerate(column_order, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thick_border
    
    # Group rows by experience
    app.logger.info('  Grouping rows by experience...')
    groups = group_by_experience(rows)
    app.logger.info(f'  Found {len(groups)} groups')
    
    # Columns that should be merged for same experience
    merge_columns = ['Associated Experience', 'Backstage Experience page', 'Public Preview Link', 'Associated Experience IDs', 'Assignee']
    merge_col_indices = [column_order.index(col) + 1 for col in merge_columns]
    
    # Write data rows
    app.logger.info(f'  Writing {len(groups)} groups to XLSX...')
    current_row = 2
    for group_idx, (experience_key, group_size, group_rows) in enumerate(groups):
        # Log progress every 100 groups
        if group_idx % 100 == 0:
            app.logger.info(f'    Processing group {group_idx}/{len(groups)}...')
        
        # Choose color for this experience group
        group_fill = colors[group_idx % len(colors)]
        
        start_row = current_row
        
        # Write all rows in this group
        for row_data in group_rows:
            for col_idx, col_name in enumerate(column_order, start=1):
                value = row_data[col_name]
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.fill = group_fill
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell.border = thin_border
                
                # Add formula for Public Preview Link column
                if col_name == 'Public Preview Link':
                    # Find the column letter for Associated Experience IDs
                    exp_ids_col_idx = column_order.index('Associated Experience IDs') + 1
                    exp_ids_col_letter = get_column_letter(exp_ids_col_idx)
                    # Set formula: ="https://app.eko.com/public/experiences/" & D2
                    cell.value = f'="https://app.eko.com/public/experiences/" & {exp_ids_col_letter}{current_row}'
                
                # Make URLs clickable (no font styling)
                elif col_name == 'Backstage Experience page' and value and (str(value).startswith('http://') or str(value).startswith('https://')):
                    cell.hyperlink = str(value)
                
            current_row += 1
        
        # Merge cells for experience identification columns if group has multiple tickets
        if group_size > 1:
            end_row = current_row - 1
            for col_idx in merge_col_indices:
                ws.merge_cells(
                    start_row=start_row,
                    start_column=col_idx,
                    end_row=end_row,
                    end_column=col_idx
                )
                # Center the merged cell content
                merged_cell = ws.cell(row=start_row, column=col_idx)
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                merged_cell.border = thick_border
                
                # Make URLs clickable (no font styling)
                col_name = column_order[col_idx - 1]
                if col_name == 'Backstage Experience page' and merged_cell.value:
                    url = str(merged_cell.value)
                    if url.startswith('http://') or url.startswith('https://'):
                        merged_cell.hyperlink = url
    
    # Auto-adjust column widths
    for col_idx, col_name in enumerate(column_order, start=1):
        col_letter = get_column_letter(col_idx)
        
        # Set minimum and maximum widths
        if col_name in ['Ticket description', 'Ticket name']:
            ws.column_dimensions[col_letter].width = 40
        elif col_name in ['Associated Experience', 'Backstage Experience page', 'Public Preview Link']:
            ws.column_dimensions[col_letter].width = 30
        elif col_name in ['Associated Experience IDs', 'Ticket ID']:
            ws.column_dimensions[col_letter].width = 20
        elif col_name in ['Assignee', 'Ticket Category', 'Sub Category', 'Invalid ticket']:
            ws.column_dimensions[col_letter].width = 20
        elif col_name in ['Additional Notes', 'Notes - Admin', 'Non-ticketed issues']:
            ws.column_dimensions[col_letter].width = 30
        elif col_name in ['Next step', 'Last Updated - Status', 'Last Updated - Next step', 'Last Updated - Invalid Ticket', 'Last Updated - Non Ticketed Issues']:
            ws.column_dimensions[col_letter].width = 25
        else:
            ws.column_dimensions[col_letter].width = 15
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    max_row = len(rows) + 1
    
    # Add data validation dropdown for Ticket status column
    status_col_idx = column_order.index('Ticket status') + 1
    status_col_letter = get_column_letter(status_col_idx)
    
    dv_status = DataValidation(
        type="list",
        formula1='"Todo,In Progress,Done,Resolved,Stuck,Other ticket stuck"',
        allow_blank=True
    )
    dv_status.error = 'Please select a valid status from the dropdown'
    dv_status.errorTitle = 'Invalid Status'
    dv_status.prompt = 'Select a status from the dropdown'
    dv_status.promptTitle = 'Ticket Status'
    dv_status.add(f'{status_col_letter}2:{status_col_letter}{max_row}')
    ws.add_data_validation(dv_status)
    
    # Add data validation dropdown for Next step column (single select)
    next_step_col_idx = column_order.index('Next step') + 1
    next_step_col_letter = get_column_letter(next_step_col_idx)
    
    dv_next_step = DataValidation(
        type="list",
        formula1='"Reshoot,Manual post"',
        allow_blank=True
    )
    dv_next_step.prompt = 'Select one option'
    dv_next_step.promptTitle = 'Next Step'
    dv_next_step.add(f'{next_step_col_letter}2:{next_step_col_letter}{max_row}')
    ws.add_data_validation(dv_next_step)
    
    # Add data validation dropdown for Assignee column (single select)
    assignee_col_idx = column_order.index('Assignee') + 1
    assignee_col_letter = get_column_letter(assignee_col_idx)
    
    dv_assignee = DataValidation(
        type="list",
        formula1=f'"{",".join(ASSIGNEES)}"',
        allow_blank=True
    )
    dv_assignee.prompt = 'Select an assignee'
    dv_assignee.promptTitle = 'Assignee'
    dv_assignee.add(f'{assignee_col_letter}2:{assignee_col_letter}{max_row}')
    ws.add_data_validation(dv_assignee)
    
    # Save workbook
    app.logger.info(f'  Saving workbook to {output_path}...')
    wb.save(output_path)
    app.logger.info('  Workbook saved successfully')
    
    return groups


def process_csv_file(input_path, output_path, assignee_distribution=None, returning_experiences=None):
    """Process CSV file and create organized XLSX output."""
    # Read CSV
    app.logger.info('Reading CSV...')
    rows = read_csv_robust(input_path)
    app.logger.info(f'Read {len(rows)} rows')
    
    # Process and reorder columns
    app.logger.info('Processing columns...')
    processed_rows, column_order = extract_and_reorder_columns(rows)
    
    # Group experiences
    app.logger.info('Grouping experiences...')
    groups = group_by_experience(processed_rows)
    app.logger.info(f'Found {len(groups)} unique experiences')
    
    # Create a dict of returning experience IDs to assignees for quick lookup
    returning_exp_map = {}
    if returning_experiences:
        for exp in returning_experiences:
            returning_exp_map[exp['experience_id']] = exp['assignee']
    
    # Auto-assign returning experiences first
    if returning_exp_map:
        for key, ticketnum, group_rows in groups:
            exp_id = str(key[2]).strip() if key[2] else ''
            if exp_id in returning_exp_map:
                # This is a returning experience - auto-assign
                assignee = returning_exp_map[exp_id]
                for row in group_rows:
                    row['Assignee'] = assignee
    
    # Assign NEW experiences to assignees if distribution is provided
    if assignee_distribution:
        app.logger.info('Assigning experiences to assignees...')
        groups = assign_experiences_to_assignees(groups, assignee_distribution, returning_exp_map)
    
    # Create XLSX (we need to flatten groups back to rows for create_xlsx)
    app.logger.info('Flattening groups...')
    all_rows = []
    for _, _, group_rows in groups:
        all_rows.extend(group_rows)
    app.logger.info(f'Total rows for XLSX: {len(all_rows)}')
    
    # Create XLSX
    app.logger.info('Creating XLSX file (this may take a while for large files)...')
    groups = create_xlsx(all_rows, column_order, output_path)
    app.logger.info('XLSX file created successfully')
    
    # Return statistics
    multi_ticket_experiences = [g for g in groups if g[1] > 1]
    stats = {
        'total_tickets': len(all_rows),
        'unique_experiences': len(groups),
        'multi_ticket_experiences': len(multi_ticket_experiences)
    }
    
    return stats


@app.route('/')
def index():
    """Render the main upload page."""
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_file():
    """Handle file upload and show assignee distribution page."""
    app.logger.info('========== UPLOAD STARTED ==========')
    if 'file' not in request.files:
        flash('No file selected', 'error')
        return redirect(url_for('index'))
    
    file = request.files['file']
    
    if file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('index'))
    
    if not allowed_file(file.filename):
        flash('Invalid file type. Please upload a CSV file.', 'error')
        return redirect(url_for('index'))
    
    try:
        # Clean up any old session data first
        old_uploaded_file = session.get('uploaded_filename')
        if old_uploaded_file:
            old_file_path = os.path.join(app.config['UPLOAD_FOLDER'], old_uploaded_file)
            if os.path.exists(old_file_path):
                os.remove(old_file_path)
        
        old_returning_file = session.get('returning_exp_file')
        if old_returning_file:
            old_returning_path = os.path.join(app.config['UPLOAD_FOLDER'], old_returning_file)
            if os.path.exists(old_returning_path):
                os.remove(old_returning_path)
        
        # Clear old session
        session.clear()
        
        # Save uploaded file with timestamp to avoid conflicts
        timestamp = int(datetime.now().timestamp() * 1000)
        original_filename = secure_filename(file.filename)
        filename = f"{timestamp}_{original_filename}"
        input_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(input_path)
        
        # Read the CSV to count experiences
        rows = read_csv_robust(input_path)
        processed_rows, _ = extract_and_reorder_columns(rows)
        groups = group_by_experience(processed_rows)
        
        # Get existing assignments from Google Sheets
        app.logger.info('Starting Google Sheets check...')
        app.logger.info(f'GOOGLE_SHEETS_ENABLED: {GOOGLE_SHEETS_ENABLED}')
        existing_assignments = get_existing_assignments_from_sheet()
        app.logger.info(f'Found {len(existing_assignments)} existing assignments in Google Sheets')
        
        # Log first 10 for debugging
        if existing_assignments:
            sample_ids = list(existing_assignments.keys())[:10]
            app.logger.info(f'Sample Experience IDs from sheet: {sample_ids}')
        
        # Log first 10 from CSV for comparison
        csv_exp_ids = [str(key[2]).strip() for key, _, _ in groups if key[2]][:10]
        app.logger.info(f'Sample Experience IDs from CSV: {csv_exp_ids}')
        
        # Separate new vs returning experiences
        new_experiences = []
        returning_experiences = []
        
        for key, _, _ in groups:
            exp_id = str(key[2]).strip() if key[2] else ''
            if exp_id and exp_id in existing_assignments:
                # This is a returning experience
                app.logger.info(f'✓ MATCH FOUND: {exp_id} -> {existing_assignments[exp_id]["assignee"]}')
                returning_experiences.append({
                    'experience_id': exp_id,
                    'experience_name': key[1],
                    'assignee': existing_assignments[exp_id]['assignee'],
                    'date_assigned': existing_assignments[exp_id].get('date_assigned', 'Unknown'),
                    'previous_tab': existing_assignments[exp_id].get('tab', 'Unknown')
                })
            elif exp_id:
                # This is a new experience
                new_experiences.append(exp_id)
        
        # Store returning experiences in a temp file (session cookie too small for large data)
        returning_exp_filename = f"{timestamp}_returning.json"
        returning_exp_path = os.path.join(app.config['UPLOAD_FOLDER'], returning_exp_filename)
        with open(returning_exp_path, 'w') as f:
            json.dump(returning_experiences, f)
        
        # Store in session
        session['uploaded_filename'] = filename
        session['original_filename'] = original_filename
        session['total_experiences'] = len(groups)
        session['new_experiences'] = len(new_experiences)
        session['returning_exp_file'] = returning_exp_filename
        
        # Log the results
        app.logger.info(f'Session after upload: {dict(session)}')
        app.logger.info(f"Upload complete: {len(new_experiences)} new, {len(returning_experiences)} returning")
        
        # Redirect to assignee distribution page
        return redirect(url_for('assignee_distribution'))
        
    except Exception as e:
        flash(f'Error processing file: {str(e)}', 'error')
        if 'input_path' in locals() and os.path.exists(input_path):
            os.remove(input_path)
        return redirect(url_for('index'))


@app.route('/assignee-distribution')
def assignee_distribution():
    """Show assignee distribution page."""
    if 'uploaded_filename' not in session:
        flash('No file uploaded', 'error')
        return redirect(url_for('index'))
    
    # Load returning experiences from temp file
    returning_experiences = []
    returning_exp_file = session.get('returning_exp_file')
    if returning_exp_file:
        returning_exp_path = os.path.join(app.config['UPLOAD_FOLDER'], returning_exp_file)
        if os.path.exists(returning_exp_path):
            with open(returning_exp_path, 'r') as f:
                returning_experiences = json.load(f)
    
    new_experiences = session.get('new_experiences', session.get('total_experiences', 0))
    
    return render_template('assignee_distribution.html', 
                          assignees=ASSIGNEES,
                          total_experiences=session.get('total_experiences', 0),
                          new_experiences=new_experiences,
                          returning_experiences=returning_experiences)


@app.route('/process', methods=['POST'])
def process_file():
    """Process the file with assignee distribution."""
    app.logger.info(f'Session keys: {list(session.keys())}')
    app.logger.info(f'Session data: {dict(session)}')
    
    if 'uploaded_filename' not in session:
        app.logger.error('No uploaded_filename in session!')
        flash('No file uploaded', 'error')
        return redirect(url_for('index'))
    
    try:
        filename = session['uploaded_filename']
        input_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        app.logger.info(f'Processing file: {filename}')
        app.logger.info(f'Input path: {input_path}')
        app.logger.info(f'Temp dir: {app.config["UPLOAD_FOLDER"]}')
        
        # List all files in temp dir
        try:
            temp_files = os.listdir(app.config['UPLOAD_FOLDER'])
            csv_files = [f for f in temp_files if f.endswith('.csv')]
            app.logger.info(f'CSV files in temp dir: {csv_files[:10]}')  # First 10
        except Exception as e:
            app.logger.error(f'Error listing temp dir: {e}')
        
        if not os.path.exists(input_path):
            app.logger.error(f'Input file not found at: {input_path}')
            flash('Uploaded file not found', 'error')
            return redirect(url_for('index'))
        
        # Get assignee distribution from request
        assignee_distribution = {}
        distribution_data = request.json if request.is_json else request.form
        
        for assignee in ASSIGNEES:
            count = int(distribution_data.get(f'assignee_{assignee}', 0))
            if count > 0:
                assignee_distribution[assignee] = count
        
        # Create output filename
        original_filename = session.get('original_filename', filename)
        output_filename = Path(original_filename).stem + '_organized.xlsx'
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
        
        app.logger.info(f'Output path: {output_path}')
        
        # Load returning experiences from temp file
        returning_experiences = []
        returning_exp_file = session.get('returning_exp_file')
        if returning_exp_file:
            returning_exp_path = os.path.join(app.config['UPLOAD_FOLDER'], returning_exp_file)
            if os.path.exists(returning_exp_path):
                with open(returning_exp_path, 'r') as f:
                    returning_experiences = json.load(f)
        
        # Process the file with assignee distribution and returning experiences
        app.logger.info('Starting process_csv_file...')
        stats = process_csv_file(input_path, output_path, assignee_distribution, returning_experiences)
        
        app.logger.info(f'File processed successfully. Stats: {stats}')
        
        # Verify output file exists
        if not os.path.exists(output_path):
            app.logger.error(f'Output file was not created at: {output_path}')
            raise Exception('Failed to create output file')
        
        app.logger.info(f'Output file size: {os.path.getsize(output_path)} bytes')
        
        # Set up cleanup after file is sent
        @after_this_request
        def cleanup(response):
            try:
                # Clean up ALL temp files after response is sent
                if os.path.exists(output_path):
                    os.remove(output_path)
                    app.logger.info(f'Cleaned up output file: {output_path}')
                if os.path.exists(input_path):
                    os.remove(input_path)
                    app.logger.info(f'Cleaned up input file: {input_path}')
                if returning_exp_file:
                    returning_exp_path_cleanup = os.path.join(app.config['UPLOAD_FOLDER'], returning_exp_file)
                    if os.path.exists(returning_exp_path_cleanup):
                        os.remove(returning_exp_path_cleanup)
                        app.logger.info(f'Cleaned up returning exp file: {returning_exp_path_cleanup}')
            except Exception as e:
                app.logger.error(f'Error during cleanup: {e}')
            return response
        
        # Clear session before sending (so user can immediately upload a new file)
        session.pop('uploaded_filename', None)
        session.pop('original_filename', None)
        session.pop('total_experiences', None)
        session.pop('new_experiences', None)
        session.pop('returning_exp_file', None)
        
        # Send the file for download
        app.logger.info(f'Sending file for download: {output_filename}')
        return send_file(
            output_path,
            as_attachment=True,
            download_name=output_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        app.logger.error(f'Error in process_file: {str(e)}', exc_info=True)
        flash(f'Error processing file: {str(e)}', 'error')
        # Clean up files if they exist
        if 'input_path' in locals() and os.path.exists(input_path):
            os.remove(input_path)
        if 'output_path' in locals() and os.path.exists(output_path):
            os.remove(output_path)
        session.pop('uploaded_filename', None)
        session.pop('total_experiences', None)
        return redirect(url_for('index'))


if __name__ == '__main__':
    # For local development
    port = int(os.environ.get('PORT', 5001))
    app.run(debug=True, host='0.0.0.0', port=port)

