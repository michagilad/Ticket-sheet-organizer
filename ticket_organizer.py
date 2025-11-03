#!/usr/bin/env python3
"""
Ticket CSV to XLSX Organizer
Processes ticket export CSV files and creates organized XLSX output with proper formatting.
"""

import csv
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Available assignees list
ASSIGNEES = [
    'Cs. Zsófi',
    'Ricsi',
    'Attila',
    'Lili',
    'Flóra',
    'Csabi',
    'Marci',
    'Gábor',
    'Bálint',
    'B.Zsófi',
    'Muki',
    'Péter',
    'Adri',
    'Era',
    'Endre',
    'Vencel',
    'Áron'
]


def read_csv_robust(file_path):
    """
    Read CSV file with robust handling of line breaks and special characters.
    """
    with open(file_path, 'r', encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    return rows


def extract_and_reorder_columns(rows):
    """
    Extract only the required columns and reorder them as specified.
    """
    # Define the columns we need from input
    input_columns = [
        'Ticket ID',
        'Backstage Experience page',
        'Company',
        'Create date',
        'Priority',
        'Record source detail 1',
        'Ticket description',
        'Ticket name',
        'Ticket status',
        'Associated Experience IDs',
        'Associated Experience',
        'Assignee',
        'Invalid ticket',
        'Ticket Category',
        'Sub Category',
        'Non-ticketed issues',
        'Additional Notes',
        'Notes - Admin',
        'Next step',
        'Last Updated - Status',
        'Last Updated - Next step'
    ]
    
    # Define the output order
    output_order = [
        'Associated Experience',
        'Backstage Experience page',
        'Associated Experience IDs',
        'Assignee',
        'Ticket name',
        'Ticket description',
        'Invalid ticket',
        'Ticket Category',
        'Ticket status',
        'Sub Category',
        'Non-ticketed issues',
        'Additional Notes',
        'Notes - Admin',
        'Next step',
        'Create date',
        'Ticket ID',
        'Record source detail 1',
        'Company',
        'Priority',
        'Last Updated - Status',
        'Last Updated - Next step'
    ]
    
    processed_rows = []
    for row in rows:
        processed_row = {}
        for col in output_order:
            processed_row[col] = row.get(col, '')
        processed_rows.append(processed_row)
    
    # Sort by Associated Experience IDs (numerically) before grouping
    # This ensures that all rows with the same Experience ID are consecutive
    def get_experience_id(row):
        exp_id = row.get('Associated Experience IDs', '')
        if not exp_id or not str(exp_id).strip():
            # Empty IDs go to the end
            return float('inf')
        try:
            # Try to convert to integer for proper numeric sorting
            return int(exp_id)
        except (ValueError, TypeError):
            # If conversion fails, put at the end
            return float('inf')
    
    processed_rows.sort(key=get_experience_id)
    
    return processed_rows, output_order


def group_by_experience(rows):
    """
    Group rows by experience to identify which rows belong to the same experience.
    Returns a list of tuples: (experience_key, group_size, rows_in_group)
    """
    groups = []
    current_group = []
    current_key = None
    
    for row in rows:
        # Create a key based on the three identifying columns
        key = (
            row.get('Associated Experience', ''),
            row.get('Backstage Experience page', ''),
            row.get('Associated Experience IDs', '')
        )
        
        # Check if key is all empty (don't group empty rows together)
        is_empty_key = all(k == '' for k in key)
        
        if current_key is None:
            # First row
            current_key = key
            current_group = [row]
        elif key == current_key and not is_empty_key:
            # Same experience AND not empty, add to current group
            current_group.append(row)
        else:
            # New experience OR empty key, save previous group and start new one
            groups.append((current_key, len(current_group), current_group))
            current_key = key
            current_group = [row]
    
    # Don't forget the last group
    if current_group:
        groups.append((current_key, len(current_group), current_group))
    
    return groups


def create_xlsx(rows, column_order, output_path):
    """
    Create XLSX file with proper formatting, merging cells for identical experiences.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Organized Tickets"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Alternating colors for experience groups
    colors = [
        PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"),
        PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
    ]
    
    # Border styles
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    thick_border = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )
    
    # Write header row
    for col_idx, col_name in enumerate(column_order, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thick_border
    
    # Group rows by experience
    groups = group_by_experience(rows)
    
    # Columns that should be merged for same experience
    merge_columns = ['Associated Experience', 'Backstage Experience page', 'Associated Experience IDs', 'Assignee']
    merge_col_indices = [column_order.index(col) + 1 for col in merge_columns]
    
    # Write data rows
    current_row = 2
    for group_idx, (experience_key, group_size, group_rows) in enumerate(groups):
        # Choose color for this experience group
        group_fill = colors[group_idx % len(colors)]
        
        start_row = current_row
        
        # Write all rows in this group
        for row_data in group_rows:
            for col_idx, col_name in enumerate(column_order, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=row_data[col_name])
                cell.fill = group_fill
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell.border = thin_border
                
                # Make Associated Experience and Ticket name bold
                if col_name in ['Associated Experience', 'Ticket name']:
                    cell.font = Font(bold=True)
                
                # Make Backstage Experience page URLs clickable
                if col_name == 'Backstage Experience page' and row_data[col_name]:
                    url = row_data[col_name]
                    if url.startswith('http://') or url.startswith('https://'):
                        cell.hyperlink = url
                        cell.font = Font(color="0563C1", underline="single")
                
            current_row += 1
        
        # Merge cells for experience identification columns if group has multiple tickets
        if group_size > 1:
            end_row = current_row - 1
            for col_idx in merge_col_indices:
                ws.merge_cells(
                    start_row=start_row,
                    start_column=col_idx,
                    end_row=end_row,
                    end_column=col_idx
                )
                # Center the merged cell content
                merged_cell = ws.cell(row=start_row, column=col_idx)
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                merged_cell.border = thick_border
                
                # Preserve bold formatting for Associated Experience
                col_name = column_order[col_idx - 1]
                if col_name == 'Associated Experience':
                    merged_cell.font = Font(bold=True)
                
                # Preserve hyperlink formatting for Backstage Experience page
                if col_name == 'Backstage Experience page' and merged_cell.value:
                    url = merged_cell.value
                    if url.startswith('http://') or url.startswith('https://'):
                        merged_cell.hyperlink = url
                        merged_cell.font = Font(color="0563C1", underline="single")
    
    # Auto-adjust column widths
    for col_idx, col_name in enumerate(column_order, start=1):
        col_letter = get_column_letter(col_idx)
        
        # Set minimum and maximum widths
        if col_name in ['Ticket description', 'Ticket name']:
            ws.column_dimensions[col_letter].width = 40
        elif col_name in ['Associated Experience', 'Backstage Experience page']:
            ws.column_dimensions[col_letter].width = 30
        elif col_name in ['Associated Experience IDs', 'Ticket ID']:
            ws.column_dimensions[col_letter].width = 20
        elif col_name in ['Assignee', 'Ticket Category', 'Sub Category', 'Invalid ticket']:
            ws.column_dimensions[col_letter].width = 20
        elif col_name in ['Additional Notes', 'Notes - Admin', 'Non-ticketed issues']:
            ws.column_dimensions[col_letter].width = 30
        elif col_name in ['Next step', 'Last Updated - Status', 'Last Updated - Next step']:
            ws.column_dimensions[col_letter].width = 25
        else:
            ws.column_dimensions[col_letter].width = 15
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    max_row = len(rows) + 1
    
    # Add data validation dropdown for Ticket status column
    status_col_idx = column_order.index('Ticket status') + 1
    status_col_letter = get_column_letter(status_col_idx)
    
    dv_status = DataValidation(
        type="list",
        formula1='"Todo,In Progress,Done,Resolved,Stuck"',
        allow_blank=True
    )
    dv_status.error = 'Please select a valid status from the dropdown'
    dv_status.errorTitle = 'Invalid Status'
    dv_status.prompt = 'Select a status from the dropdown'
    dv_status.promptTitle = 'Ticket Status'
    dv_status.add(f'{status_col_letter}2:{status_col_letter}{max_row}')
    ws.add_data_validation(dv_status)
    
    # Add data validation dropdown for Next step column (single select)
    next_step_col_idx = column_order.index('Next step') + 1
    next_step_col_letter = get_column_letter(next_step_col_idx)
    
    dv_next_step = DataValidation(
        type="list",
        formula1='"Reshoot,Manual post"',
        allow_blank=True
    )
    dv_next_step.prompt = 'Select one option'
    dv_next_step.promptTitle = 'Next Step'
    dv_next_step.add(f'{next_step_col_letter}2:{next_step_col_letter}{max_row}')
    ws.add_data_validation(dv_next_step)
    
    # Add data validation dropdown for Assignee column (single select)
    assignee_col_idx = column_order.index('Assignee') + 1
    assignee_col_letter = get_column_letter(assignee_col_idx)
    
    dv_assignee = DataValidation(
        type="list",
        formula1=f'"{",".join(ASSIGNEES)}"',
        allow_blank=True
    )
    dv_assignee.prompt = 'Select an assignee'
    dv_assignee.promptTitle = 'Assignee'
    dv_assignee.add(f'{assignee_col_letter}2:{assignee_col_letter}{max_row}')
    ws.add_data_validation(dv_assignee)
    
    # Save workbook
    wb.save(output_path)


def main():
    """
    Main function to process CSV and create XLSX output.
    """
    if len(sys.argv) < 2:
        print("Usage: python ticket_organizer.py <input_csv_file> [output_xlsx_file]")
        print("\nExample:")
        print("  python ticket_organizer.py tickets.csv")
        print("  python ticket_organizer.py tickets.csv organized_tickets.xlsx")
        sys.exit(1)
    
    input_file = sys.argv[1]
    
    # Generate output filename if not provided
    if len(sys.argv) >= 3:
        output_file = sys.argv[2]
    else:
        input_path = Path(input_file)
        output_file = input_path.stem + "_organized.xlsx"
    
    # Validate input file exists
    if not Path(input_file).exists():
        print(f"Error: Input file '{input_file}' not found.")
        sys.exit(1)
    
    print(f"Reading CSV file: {input_file}")
    try:
        rows = read_csv_robust(input_file)
        print(f"  Found {len(rows)} tickets")
    except Exception as e:
        print(f"Error reading CSV file: {e}")
        sys.exit(1)
    
    print("Processing and reordering columns...")
    try:
        processed_rows, column_order = extract_and_reorder_columns(rows)
    except Exception as e:
        print(f"Error processing columns: {e}")
        sys.exit(1)
    
    print(f"Creating XLSX file: {output_file}")
    try:
        create_xlsx(processed_rows, column_order, output_file)
        print(f"✓ Successfully created {output_file}")
        
        # Print summary
        groups = group_by_experience(processed_rows)
        multi_ticket_experiences = [g for g in groups if g[1] > 1]
        print(f"\nSummary:")
        print(f"  Total tickets: {len(processed_rows)}")
        print(f"  Unique experiences: {len(groups)}")
        print(f"  Experiences with multiple tickets: {len(multi_ticket_experiences)}")
        
    except Exception as e:
        print(f"Error creating XLSX file: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()


